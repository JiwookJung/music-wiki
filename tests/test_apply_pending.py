"""ser8 발급 → 엑셀 반영(apply_pending) 동작."""
import importlib.util
import json
import sys
from pathlib import Path

import openpyxl

SCRIPT = Path(__file__).resolve().parents[1] / "inventory" / "scripts" / "apply_pending.py"


def _load(tmp_path, xlsx, pending):
    spec = importlib.util.spec_from_file_location("ap", SCRIPT)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    mod.XLSX = str(xlsx)
    mod.PENDING = str(pending)
    mod.ARCHIVE = str(tmp_path / "applied.json")
    return mod


def _xlsx(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LP중앙선반2열1층"
    ws.append(["순번", "분류코드", "매체", "장르", "아티스트", "앨범제목", "대표곡",
               "작곡가", "연주자", "레이블/카탈로그번호", "비고"])
    ws.append([1, "J-B01-01", "LP", "재즈", "Bill Evans", "Waltz", "", "", "", "", ""])
    p = tmp_path / "inv.xlsx"
    wb.save(p)
    return p


def test_pending_row_is_appended(tmp_path, monkeypatch):
    xlsx = _xlsx(tmp_path)
    pending = tmp_path / "pending.json"
    pending.write_text(json.dumps([{
        "sheet": "LP중앙선반2열1층", "medium": "LP", "genre": "재즈",
        "artist": "Keith Jarrett", "album": "The Köln Concert",
        "code": "J-K01-02", "notes": "신규 구매"}], ensure_ascii=False), encoding="utf-8")
    mod = _load(tmp_path, xlsx, pending)
    monkeypatch.setattr(sys, "argv", ["apply_pending.py"])
    mod.main()
    ws = openpyxl.load_workbook(xlsx)["LP중앙선반2열1층"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    added = rows[-1]
    assert added[4] == "Keith Jarrett" and added[5] == "The Köln Concert"
    assert added[1] == "J-K01-02"          # 분류코드 보존
    assert "웹등록" in str(added[10])       # 비고 스탬프
    assert not pending.exists()             # 큐 비움


def test_duplicate_is_skipped_and_queue_cleared(tmp_path, monkeypatch):
    xlsx = _xlsx(tmp_path)
    pending = tmp_path / "pending.json"
    pending.write_text(json.dumps([{
        "sheet": "LP중앙선반2열1층", "medium": "LP", "genre": "재즈",
        "artist": "Bill Evans", "album": "Waltz", "code": "J-B01-01"}],
        ensure_ascii=False), encoding="utf-8")
    mod = _load(tmp_path, xlsx, pending)
    monkeypatch.setattr(sys, "argv", ["apply_pending.py"])
    mod.main()
    ws = openpyxl.load_workbook(xlsx)["LP중앙선반2열1층"]
    assert len(list(ws.iter_rows(min_row=2, values_only=True))) == 1   # 중복 미추가
    assert not pending.exists()


def test_dry_run_changes_nothing(tmp_path, monkeypatch):
    xlsx = _xlsx(tmp_path)
    pending = tmp_path / "pending.json"
    payload = [{"sheet": "LP중앙선반2열1층", "medium": "LP", "genre": "재즈",
                "artist": "New Artist", "album": "New Album", "code": "J-N01-01"}]
    pending.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    mod = _load(tmp_path, xlsx, pending)
    monkeypatch.setattr(sys, "argv", ["apply_pending.py", "--dry-run"])
    mod.main()
    ws = openpyxl.load_workbook(xlsx)["LP중앙선반2열1층"]
    assert len(list(ws.iter_rows(min_row=2, values_only=True))) == 1
    assert json.loads(pending.read_text(encoding="utf-8")) == payload   # 큐 유지
