#!/usr/bin/env python3
"""디지털 라이브러리(music-wiki.db) ↔ 실물 인벤토리(LP/CD 분류코드) 연동.

- 매칭: 정규화된 아티스트+앨범 정확 일치(관사 제거 포함).
- 산출:
  1) inventory/data/digital_links.json  → pipeline.py 가 엑셀 '디지털' 열에 ✓ 표시
  2) music-wiki.db album.physical_code  → 위키 앨범 페이지에 "실물 음반: LP J-B01-02" 표시
     (LP 코드는 has_vinyl=1 배지도 켜짐)
- 재실행 안전(멱등). 위키 반영은 `music-wiki build-wiki`(또는 --wiki 옵션)로.
"""
import argparse
import json
import os
import re
import sys
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
INV = os.path.dirname(HERE)
REPO = os.path.dirname(INV)
sys.path.insert(0, os.path.join(REPO, "src"))

XLSX = "/home/neotango/lpcd_image/20260726_LPCD_목록_v2.xlsx"
if not os.path.exists(XLSX):
    XLSX = os.path.join(INV, "20260726_LPCD_목록_v2.xlsx")
DB = os.path.expanduser("~/music-wiki-vault/music-wiki.db")
LINKS = os.path.join(INV, "data", "digital_links.json")
SUMMARY_SHEETS = {"분류코드표", "중복목록", "미식별목록", "정리계획", "라벨인쇄"}


def norm(s):
    return re.sub(r"[^a-z0-9가-힣]", "", str(s or "").lower())


def artist_key(name):
    s = str(name or "")
    s = re.sub(r"^\s*the\s+", "", s, flags=re.I)
    s = re.sub(r"^\s*el\s+", "", s, flags=re.I)
    if re.match(r"^\s*los\s+(?!angeles\b)", s, flags=re.I):
        s = re.sub(r"^\s*los\s+", "", s, flags=re.I)
    return s


def strip_paren(s):
    return re.sub(r"\s*[\(（][^\)）]*[\)）]", "", str(s or "")).strip()


def akeys(artist):
    """아티스트 매칭 후보키(원형·관사제거·괄호제거)."""
    out = set()
    for v in (artist, artist_key(artist), strip_paren(artist), artist_key(strip_paren(artist))):
        n = norm(v)
        if n:
            out.add(n)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--wiki", action="store_true", help="연동 후 마크다운 위키 재생성")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    # 실물 인벤토리: (artist_norm후보, album_norm) -> {code, media}
    wb = openpyxl.load_workbook(XLSX)
    phys = {}
    media = defaultdict(set)
    for name in wb.sheetnames:
        if name in SUMMARY_SHEETS:
            continue
        ws = wb[name]
        hdr = [c.value for c in ws[1]]
        ix = {h: hdr.index(h) for h in ("분류코드", "매체", "아티스트", "앨범제목")}
        for r in ws.iter_rows(min_row=2, values_only=True):
            code = str(r[ix["분류코드"]] or "").strip()
            a, al = str(r[ix["아티스트"]] or ""), str(r[ix["앨범제목"]] or "")
            if not (code and a.strip() and al.strip()):
                continue
            aln = norm(strip_paren(al)) or norm(al)
            for an in akeys(a):
                phys[(an, aln)] = code
                media[code].add(str(r[ix["매체"]] or "").strip() or "?")

    from music_wiki.core.store import Store
    s = Store.open(DB)
    s.init_schema()

    matched = []
    for artist in s.iter_artists():
        d_akeys = akeys(artist.name)
        for album in s.albums_for_artist(artist.id):
            aln = norm(strip_paren(album.title)) or norm(album.title)
            code = next((phys[(an, aln)] for an in d_akeys if (an, aln) in phys), None)
            if code:
                m = "·".join(sorted(media[code]))
                matched.append((album.id, artist.name, album.title, code, m))

    print(f"실물 코드 보유 앨범(고유): {len(set(phys.values()))} | 디지털 앨범: 1185급"
          f" | 매칭: {len(matched)}건")
    for _, a, al, code, m in matched[:10]:
        print(f"  {code:<16} [{m}] {a[:20]} — {al[:34]}")
    if len(matched) > 10:
        print(f"  … 외 {len(matched)-10}건")

    if args.dry_run:
        print("dry-run — 미반영")
        return

    links = {}
    for _, a, al, code, m in matched:
        links[f"{norm(artist_key(strip_paren(a)))}|{norm(strip_paren(al)) or norm(al)}"] = \
            {"code": code, "media": m}
    # 엑셀 ✓용: 인벤토리 쪽 키(원형 포함)도 모두 기록
    for (an, aln), code in phys.items():
        k = f"{an}|{aln}"
        if any(v["code"] == code for v in links.values()):
            links.setdefault(k, {"code": code, "media": "·".join(sorted(media[code]))})
    json.dump(links, open(LINKS, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

    for aid, _, _, code, m in matched:
        s.set_physical_code(aid, f"{m} {code}")
    print(f"DB 반영 {len(matched)}건 + digital_links.json 기록")

    if args.wiki:
        from music_wiki.core.wiki import WikiGenerator
        WikiGenerator(s).generate(os.path.expanduser("~/music-wiki-vault"))
        print("위키 재생성 완료")


if __name__ == "__main__":
    main()
