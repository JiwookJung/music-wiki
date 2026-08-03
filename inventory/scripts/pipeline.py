#!/usr/bin/env python3
"""LP/CD 인벤토리 파이프라인 — 앨범 추가/수정 후 이것 하나만 실행하면 됨.

사용법:
    1) 엑셀(20260726_LPCD_목록_v2.xlsx)의 위치 시트에 행 추가/수정
       (매체·장르·아티스트·앨범제목 필수, 클래식은 작곡가·연주자까지)
    2) python pipeline.py            # 코드 부여 + 전 시트 재생성 + 신규코드 출력
       python pipeline.py --dry-run  # 어떤 코드가 새로 나올지 미리보기만

핵심: code_registry.json 에 부여된 번호를 고정 저장 → 재실행해도 기존
분류코드는 절대 바뀌지 않고(라벨 유효), 새 아티스트는 이웃 사이 빈 번호에,
새 앨범은 다음 번호에 삽입된다. digital_links.json 이 있으면 '디지털' 열 표시.
"""
import argparse
import json
import os
import re
import unicodedata
from collections import defaultdict
from copy import copy

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
INV = os.path.dirname(HERE)
XLSX = "/home/neotango/lpcd_image/20260726_LPCD_목록_v2.xlsx"   # 사용자 작업본(정본)
if not os.path.exists(XLSX):
    XLSX = os.path.join(INV, "20260726_LPCD_목록_v2.xlsx")       # 저장소 사본(폴백)
REGISTRY = os.path.join(INV, "data", "code_registry.json")
DIGITAL = os.path.join(INV, "data", "digital_links.json")

GENRE_LETTER = {"클래식": "C", "클래식기타": "G", "재즈": "J", "탱고": "T",
                "월드": "W", "팝": "P", "가요": "K", "OST·경음악": "O",
                "경음악": "I"}          # I = instrumental. OST 와 분리된 독립 장르
# 엑셀 표기 흔들림 흡수 → 정규 장르명 (X 코드 방지)
# 주의: '경음악' 은 더 이상 OST 로 접지 않는다(독립 장르 I)
GENRE_ALIAS = {"OST": "OST·경음악", "OST/경음악": "OST·경음악",
               "OST,경음악": "OST·경음악", "사운드트랙": "OST·경음악",
               "이지리스닝": "경음악", "뉴에이지": "경음악",
               "클래식 기타": "클래식기타", "월드뮤직": "월드", "제3세계": "월드",
               "가요(한국)": "가요", "케이팝": "가요", "팝송": "팝"}


def canon_genre(g: str) -> str:
    g = str(g or "").strip()
    return GENRE_ALIAS.get(g, g)


SHEET_ORDER = ["CD윗층", "CD아래층",
               "LP중앙선반1열1층", "LP중앙선반1열2층", "LP중앙선반1열3층",
               "LP중앙선반2열1층", "LP중앙선반2열2층", "LP중앙선반2열3층",
               "LP좌측선반1층", "LP좌측선반2층", "LP좌측선반3층",
               "LP우측선반1층", "LP우측선반2층", "LP보관박스1", "LP보관박스2", "LP침대옆"]
MAIN = {"LP중앙선반1열1층", "LP중앙선반1열2층", "LP중앙선반1열3층", "LP중앙선반2열1층",
        "LP중앙선반2열2층", "LP중앙선반2열3층", "LP좌측선반1층", "LP좌측선반2층", "LP좌측선반3층"}
CAPACITY = {"LP중앙선반1열1층": 79, "LP중앙선반1열2층": 55, "LP중앙선반1열3층": 59,
            "LP중앙선반2열1층": 80, "LP중앙선반2열2층": 58, "LP중앙선반2열3층": 86,
            "LP좌측선반1층": 46, "LP좌측선반2층": 61, "LP좌측선반3층": 65}
SUMMARY_SHEETS = {"분류코드표", "중복목록", "미식별목록", "정리계획",
                  "라벨인쇄", "라벨인쇄-LP", "라벨인쇄-CD", "라벨인쇄-전체"}

CHO = "ㄱㄲㄴㄷㄸㄹㅁㅂㅃㅅㅆㅇㅈㅉㅊㅋㅌㅍㅎ"
CHO_BASE = {"ㄲ": "ㄱ", "ㄸ": "ㄷ", "ㅃ": "ㅂ", "ㅆ": "ㅅ", "ㅉ": "ㅈ"}


# 귀족 전치사·연결어는 성으로 보지 않는다 (von Karajan → Karajan)
_PARTICLES = {"von", "van", "de", "del", "della", "di", "da", "der", "den",
              "du", "la", "le", "los", "dos", "ten", "ter"}


def deaccent(s: str) -> str:
    """Dvořák → Dvorak. 악센트를 떼지 않으면 단어 분리가 깨져 성을 잘못 잡는다
    (예전 버그: 'Dvořák' → ['Dvo','k'] → 이니셜 K)."""
    # NFD 는 한글 음절도 자모로 분해하므로 반드시 NFC 로 되돌린다
    # (안 그러면 '조성우' 가 자모열이 되어 초성 판정이 X 로 떨어진다)
    stripped = "".join(c for c in unicodedata.normalize("NFD", str(s or ""))
                       if not unicodedata.combining(c))
    return unicodedata.normalize("NFC", stripped)


def name_words(name) -> list:
    """이름을 성 판정용 단어로 쪼갠다(악센트 제거 + 전치사 제외)."""
    return [w for w in re.findall(r"[A-Za-z가-힣]+", deaccent(name))
            if w.lower() not in _PARTICLES]


def initial(name):
    for ch in deaccent(name).strip():
        if "가" <= ch <= "힣":
            c = CHO[(ord(ch) - 0xAC00) // 588]
            return CHO_BASE.get(c, c)
        if ch.isascii() and ch.isalpha():
            return ch.upper()
        if ch.isdigit():
            return "0"
    return "X"


def sort_source(r) -> str:
    """분류코드의 '아티스트 자리'에 쓸 이름.

    OST 는 작곡가가 아니라 작품(영화·드라마) 제목으로 찾는 게 자연스럽다.
    엑셀 '정렬명' 열이 채워져 있으면 그것을, 비어 있으면 아티스트를 쓴다.
    """
    s = str(r.get("sort_name") or "").strip()
    return s or str(r.get("artist") or "")


def artist_key(name):
    """관사 제거 정렬형: The Beatles→Beatles, El/Los 제거(Los Angeles·I Musici 예외)."""
    s = str(name or "")
    s = re.sub(r"^\s*the\s+", "", s, flags=re.I)
    s = re.sub(r"^\s*el\s+", "", s, flags=re.I)
    if re.match(r"^\s*los\s+(?!angeles\b)", s, flags=re.I):
        s = re.sub(r"^\s*los\s+", "", s, flags=re.I)
    return s


def norm(s):
    return re.sub(r"[^a-z0-9가-힣]", "", str(s or "").lower())


# 뒤에 붙으면 버리는 단체·악기 표기. 앞자리(대표 아티스트)에는 적용하지 않는다.
_ENS_WORDS = (r"orchestra|orchestre|orkest|orchester|philharmoni\w*|symphon\w*|sinfoni\w*|"
              r"ensemble|academy|camerata|kapelle|orquesta|collegium|consort|choir|chorus|"
              r"chamber\s+\w+|his\s+\w+|오케스트라|필하모\w*|교향악단|합창단|앙상블")
_ENS = re.compile(rf"^(the\s+)?[\w\s'’.·&-]*\b({_ENS_WORDS})\b", re.I)
_INSTR = re.compile(r"^\(?\s*(guitare?|piano|cello|violin|flute|vocal|"
                    r"기타|피아노|첼로|바이올린)\s*\)?$", re.I)


def _has_han(s):
    return bool(re.search(r"[가-힣]", s))


def _has_lat(s):
    return bool(re.search(r"[A-Za-z]", s))


def _pick(a, b, korean):
    """병기된 두 표기 중 정본을 고른다. 가요는 한글, 그 밖의 장르는 로마자."""
    if not (a and b):
        return a or b
    return (a if _has_han(a) else b) if korean else (a if _has_lat(a) else b)


def lead_artist(name, genre="") -> str:
    """분류코드에 쓸 '대표 아티스트' 표기.

    ① 뒤따르는 단체명·악기표기는 버린다 — 'Andre Previn and the LSO' → 'Andre Previn'.
       뒤가 사람이면 동격 연주자로 보고 남긴다 — 'Julian Bream, John Williams' 는 그대로.
    ② 괄호 병기와 '이름 = Romanized' 를 한쪽으로 정리한다. 가요는 한글이 정본이고
       ('조용필(Cho Yong Pil)' → '조용필'), 그 밖의 장르는 로마자가 정본이다
       ('임윤찬(Yunchan Lim)' → 'Yunchan Lim'). 표기가 갈리면 같은 사람이 코드를
       두 개 받아 선반 두 곳으로 흩어진다.
    """
    s = str(name or "").strip()
    if not s:
        return s
    korean = str(genre or "").startswith("가요")
    parts = [p.strip() for p in re.split(r"\s*[,/;]\s*|\s*&\s*|\s+and\s+|\s+with\s+", s)
             if p.strip()]
    keep = []
    for p in parts:
        if keep and (_ENS.match(p) or _INSTR.match(p)):
            continue
        keep.append(p)
    # '이승환, Lee Seung-Hwan' 처럼 한글 이름과 그 로마자가 나란히 온 경우도 병기다
    if len(keep) == 2 and _has_han(keep[0]) and not _has_lat(keep[0]) \
            and _has_lat(keep[1]) and not _has_han(keep[1]):
        keep = [_pick(keep[0], keep[1], korean)]
    head = keep[0]
    m = re.match(r"^(.*?)\s*[\(（]([^)）]*)[\)）]\s*(.*)$", head)
    if m:
        head = (_pick(m.group(1).strip(), m.group(2).strip(), korean)
                + (" " + m.group(3).strip() if m.group(3).strip() else "")).strip()
    if "=" in head:
        a, b = (x.strip() for x in head.split("=", 1))
        head = _pick(a, b, korean)
    keep[0] = head
    return ", ".join(keep)


def primary(name) -> str:
    """공동 크레딧 중 맨 앞(대표) 한 사람만. 클래식 연주자 이니셜용."""
    return str(name or "").split(",")[0].strip()


def is_location_sheet(name: str) -> bool:
    """데이터 시트 = 실제 선반뿐. 파생·수작업 시트(배치·라벨·정리)는 읽지 않는다.
    이 가드가 없으면 배치-* 시트가 실물 재고로 두 번 계산돼 라벨 수가 틀어진다."""
    return name not in SUMMARY_SHEETS and re.match(r"^(LP|CD)", name) is not None


def album_key(r):
    """앨범 동일성 판정 키. 대표 아티스트 기준이라 표기가 갈려도 한 앨범으로 모인다.
    assign_codes 와 build 가 반드시 같은 키를 써야 한다(다르면 코드를 못 찾아 행이 샌다)."""
    return (norm(lead_artist(r.get("artist"), r.get("genre"))), norm(r.get("album")))


def surname_key(composer):
    words = name_words(composer)
    return words[-1].lower() if words else norm(composer)


def surname_initial(name):
    s = str(name or "").strip()
    if any("가" <= ch <= "힣" for ch in s):
        return initial(s)
    words = name_words(s)
    return words[-1][0].upper() if words else initial(s)


# ── 1. 엑셀 추출 ──────────────────────────────────────────────────────────
KEYMAP = [("order", "순번"), ("medium", "매체"), ("genre", "장르"), ("artist", "아티스트"),
          ("album", "앨범제목"), ("sort_name", "정렬명"), ("rep", "대표곡"), ("composer", "작곡가"), ("performer", "연주자"),
          ("label_cat", "레이블/카탈로그번호"), ("discogs_id", "Discogs ID"),
          ("mbid", "MusicBrainz MBID"), ("db_url", "DB링크"), ("desc", "해설"), ("notes", "비고")]


def extract(path):
    wb = openpyxl.load_workbook(path)
    rows = []
    for name in wb.sheetnames:
        if not is_location_sheet(name):
            continue
        ws = wb[name]
        hdr = [str(c.value) if c.value is not None else "" for c in ws[1]]
        ix = {h: hdr.index(h) for h in [k[1] for k in KEYMAP] if h in hdr}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or all(v is None or str(v).strip() == "" for v in r):
                continue
            row = {"sheet": name}
            for k, h in KEYMAP:
                v = r[ix[h]] if h in ix else None
                row[k] = "" if v is None else (v if isinstance(v, (int, float)) else str(v).strip())
            rows.append(row)
    return rows


# ── 2. 안정적 코드 부여 (레지스트리) ─────────────────────────────────────
def load_registry():
    if os.path.exists(REGISTRY):
        return json.load(open(REGISTRY, encoding="utf-8"))
    return {"artists": {}, "albums": {}, "composers": {}, "performers": {},
            "perf_albums": {}}


def insert_number(existing: dict, sort_key: str, all_sorted: list) -> tuple[int, bool]:
    """이웃 사이 빈 정수 번호를 찾는다. (번호, 순서보장여부) 반환."""
    used = set(existing.values())
    pos = all_sorted.index(sort_key)
    prev_no = 0
    for k in reversed(all_sorted[:pos]):
        if k in existing:
            prev_no = existing[k]
            break
    next_no = None
    for k in all_sorted[pos + 1:]:
        if k in existing:
            next_no = existing[k]
            break
    if next_no is None:                      # 맨 뒤 → prev+4 (여유 간격)
        n = prev_no + 4
        while n in used:
            n += 1
        return n, True
    for n in range(prev_no + 1, next_no):    # 사이 빈 번호
        if n not in used:
            return n, True
    n = max(used, default=0) + 1             # 빈 칸 없음 → 뒤에 붙이고 경고
    while n in used:
        n += 1
    return n, False


def media_prefix(idxs, rows) -> str:
    """같은 앨범의 매체로 접두 결정. LP 보유 시 'LP-', 아니면 'CD-'(그 외 '')."""
    media = {str(rows[j].get("medium") or "").strip().upper() for j in idxs}
    if "LP" in media:
        return "LP-"
    if "CD" in media:
        return "CD-"
    return ""


def assign_codes(rows, reg, warnings):
    albums = defaultdict(list)
    for i, r in enumerate(rows):
        if str(r.get("artist") or "").strip() and str(r.get("album") or "").strip():
            albums[album_key(r)].append(i)

    # 분류 대상 구조화
    meta = {}          # key -> ("N", letter, artist_norm, disp) | ("C", pref, ck, perf_norm, ...)
    for key, idxs in albums.items():
        r = rows[idxs[0]]
        g = canon_genre(r.get("genre"))
        letter = GENRE_LETTER.get(g, "X")
        if letter == "C" and (r.get("composer") or r.get("performer")):
            pref = "CG" if any("2864" in str(rows[j].get("label_cat") or "")
                               or rows[j]["sheet"] == "LP좌측선반3층" for j in idxs) else "C"
            comp = str(r.get("composer") or "VA").strip() or "VA"
            perf = lead_artist(r.get("performer") or r.get("artist"), g)
            meta[key] = ("C", pref, comp, perf)
        else:
            if letter == "K" and any(
                    "비매품" in (str(rows[j].get("album") or "") + str(rows[j].get("artist") or "")
                               + str(rows[j].get("notes") or "")) for j in idxs):
                letter = "KN"
            # OST 는 정렬명(작품 제목)이 우선. 비어 있을 때만 대표 아티스트를 쓴다.
            nm = str(r.get("sort_name") or "").strip() or lead_artist(r.get("artist"), g)
            meta[key] = ("N", letter, norm(artist_key(nm)), nm)

    album_code = {}
    # 일반(비클래식)
    groups = defaultdict(dict)   # (letter, init) -> {artist_norm: disp}
    for key, m in meta.items():
        if m[0] == "N":
            groups[(m[1], initial(artist_key(m[3])))][m[2]] = m[3]
    for (letter, init), members in groups.items():
        gkey = f"{letter}|{init}"
        stored = reg["artists"].setdefault(gkey, {})
        order = sorted(members, key=lambda k: norm(artist_key(members[k])))
        for ak in order:
            if ak not in stored:
                no, ok = insert_number(stored, ak, order)
                stored[ak] = no
                if not ok:
                    warnings.append(f"[순서초과] {letter}-{init} 그룹에 '{members[ak]}' → "
                                    f"{no:02d} (빈 번호 없음, 정렬순서 벗어남)")
    for key, m in meta.items():
        if m[0] != "N":
            continue
        letter, an, disp = m[1], m[2], m[3]
        gkey = f"{letter}|{initial(artist_key(disp))}"
        ano = reg["artists"][gkey][an]
        akey = f"{gkey}|{an}"
        stored_albums = reg["albums"].setdefault(akey, {})
        if key[1] not in stored_albums:
            stored_albums[key[1]] = max(stored_albums.values(), default=0) + 1
        album_code[key] = (f"{media_prefix(albums[key], rows)}{letter}-"
                           f"{initial(artist_key(disp))}{ano:02d}-"
                           f"{stored_albums[key[1]]:02d}")

    # 클래식
    cgroups = defaultdict(dict)  # (pref, cinit) -> {ck: comp_disp}
    for key, m in meta.items():
        if m[0] == "C":
            pref, comp = m[1], m[2]
            cinit = "V" if comp == "VA" else surname_initial(comp)
            cgroups[(pref, cinit)][surname_key(comp) or "va"] = comp
    for (pref, cinit), members in cgroups.items():
        gkey = f"{pref}|{cinit}"
        stored = reg["composers"].setdefault(gkey, {})
        for ck in sorted(members):
            if ck not in stored:
                used = set(stored.values())
                n = next((x for x in range(10) if x not in used), None)
                if n is None:
                    n = max(used) + 1
                    warnings.append(f"[작곡가 한자리 초과] {pref}-{cinit}: '{members[ck]}' → {n}")
                stored[ck] = n
    for key, m in meta.items():
        if m[0] != "C":
            continue
        pref, comp, perf = m[1], m[2], m[3]
        cinit = "V" if comp == "VA" else surname_initial(comp)
        ck = surname_key(comp) or "va"
        cno = reg["composers"][f"{pref}|{cinit}"][ck]
        pinit = surname_initial(primary(perf))   # 공동 크레딧은 맨 앞 사람 기준
        pgkey = f"{pref}|{ck}|{pinit}"
        stored_p = reg["performers"].setdefault(pgkey, {})
        # 협연 지휘자·오케스트라가 붙었다고 별도 연주자가 되면 안 된다.
        # 코드 이니셜이 대표 연주자 기준이므로 번호도 대표 연주자로 묶는다.
        pn = norm(primary(perf)) or "unknown"
        if pn not in stored_p:
            peers = sorted(set(list(stored_p) + [pn]))
            no, ok = insert_number(stored_p, pn, peers)
            stored_p[pn] = no
            if not ok:
                warnings.append(f"[순서초과] {pref}-{cinit}{cno} 연주자 '{perf}' → {no:02d}")
        pno = stored_p[pn]
        pakey = f"{pgkey}|{pn}"
        stored_albums = reg["perf_albums"].setdefault(pakey, {})
        if key[1] not in stored_albums:
            stored_albums[key[1]] = max(stored_albums.values(), default=0) + 1
        album_code[key] = (f"{media_prefix(albums[key], rows)}{pref}-{cinit}{cno}-"
                           f"{pinit}{pno:02d}-{stored_albums[key[1]]:02d}")

    return albums, album_code, meta


def seed_registry_from_existing(rows, reg):
    """기존 엑셀의 분류코드를 레지스트리에 1회 이식(라벨 유효성 보존)."""
    wb = openpyxl.load_workbook(XLSX)
    pat_n = re.compile(r"^(?:(?:LP|CD)-)?([A-Z]{1,2})-(.)(\d{2})-(\d{2})$")
    pat_c = re.compile(r"^(?:(?:LP|CD)-)?(C|CG)-(.)(\d)-(.)(\d{2})-(\d{2})$")
    for name in wb.sheetnames:
        if name in SUMMARY_SHEETS:
            continue
        ws = wb[name]
        hdr = [c.value for c in ws[1]]
        if "분류코드" not in hdr:
            continue
        ix = {h: hdr.index(h) for h in ("분류코드", "아티스트", "앨범제목", "작곡가", "연주자")}
        for r in ws.iter_rows(min_row=2, values_only=True):
            code = str(r[ix["분류코드"]] or "").strip()
            a = str(r[ix["아티스트"]] or "").strip()
            al = str(r[ix["앨범제목"]] or "").strip()
            if not (code and a and al):
                continue
            m = pat_c.match(code)
            if m:
                pref, cinit, cno, pinit, pno, ano = m.groups()
                comp = str(r[ix["작곡가"]] or "VA").strip() or "VA"
                perf = str(r[ix["연주자"]] or a).strip()
                ck = surname_key(comp) or "va"
                reg["composers"].setdefault(f"{pref}|{cinit}", {})[ck] = int(cno)
                pn = norm(perf) or "unknown"
                reg["performers"].setdefault(f"{pref}|{ck}|{pinit}", {})[pn] = int(pno)
                reg["perf_albums"].setdefault(f"{pref}|{ck}|{pinit}|{pn}", {})[norm(al)] = int(ano)
                continue
            m = pat_n.match(code)
            if m:
                letter, init, ano_, alno = m.groups()
                an = norm(artist_key(a))
                reg["artists"].setdefault(f"{letter}|{init}", {})[an] = int(ano_)
                reg["albums"].setdefault(f"{letter}|{init}|{an}", {})[norm(al)] = int(alno)


# ── 3. 워크북 생성 ────────────────────────────────────────────────────────
HDR = ["순번", "분류코드", "매체", "장르", "아티스트", "앨범제목", "정렬명", "대표곡",
       "작곡가", "연주자", "레이블/카탈로그번호", "Discogs ID", "MusicBrainz MBID",
       "DB링크", "디지털", "중복/이동", "해설", "비고"]
KEYS = ["order", "code", "medium", "genre", "artist", "album", "sort_name", "rep",
        "composer", "performer", "label_cat", "discogs_id", "mbid",
        "db_url", "digital", "dup", "desc", "notes"]
WID = [6, 18, 6, 10, 24, 32, 26, 26, 14, 16, 20, 12, 26, 38, 8, 22, 44, 22]
CENTER = {"순번", "분류코드", "매체", "장르", "Discogs ID", "디지털", "매수", "엑셀행"}
HF, HFo = PatternFill("solid", fgColor="305496"), Font(bold=True, color="FFFFFF")


def head(ws, headers, widths=None):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HF, HFo
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if widths:
            ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    ws.freeze_panes = "A2"


def blankv(v):
    return not str(v or "").strip()


# ── 라벨 섹션 ────────────────────────────────────────────────────────────
# 섹션 = 선반 위에서 칸막이로 나뉘는 묶음. 라벨 윗줄에 찍어 어디에 꽂을지 바로 안다.
SECTION_NAME = {"클래식": "클래식", "클래식기타": "클래식기타", "경음악": "경음악",
                "재즈": "재즈", "가요": "가요", "OST·경음악": "OST", "팝": "팝",
                "탱고": "탱고", "월드": "월드"}
BIG_PERF_MIN = 3          # 이 장수 이상인 연주자는 작곡가와 별도 섹션을 갖는다
# 연주자 칸에 작곡가명·모음집이 들어간 행은 연주자 섹션 대상이 아니다
_NOT_PERF = {"various", "va", "bach", "beethoven", "mozart", "schubert", "chopin",
             "brahms", "vivaldi", "handel", "haydn", "verdi", "schumann"}


def big_performers(rows, albums, album_code, minimum=BIG_PERF_MIN):
    """작곡가와 따로 모을 만한 클래식 연주자를 C/CG 계열별로 센다.

    C(일반 클래식)와 CG(그라모폰)는 서로 다른 선반에 있으므로 기준을 함께 세면
    한쪽 선반에 그 연주자 음반이 한 장뿐인데도 혼자 한 섹션을 차지하게 된다.
    """
    cnt = {"C": defaultdict(int), "CG": defaultdict(int)}
    for key, idxs in albums.items():
        r = rows[idxs[0]]
        if canon_genre(r.get("genre")) != "클래식":
            continue
        # 섹션은 LP 선반을 나누는 구분판이다. CD 만 있는 음반은 세지 않는다.
        if not any(str(rows[i].get("medium") or "").strip().upper() == "LP" for i in idxs):
            continue
        pref = "CG" if re.sub(r"^(?:LP|CD)-", "",
                              album_code.get(key, "")).startswith("CG") else "C"
        pn = primary(lead_artist(r.get("performer") or r.get("artist"), "클래식"))
        if norm(pn) and norm(pn) not in _NOT_PERF:
            cnt[pref][pn] += 1
    return {k: {n for n, v in c.items() if v >= minimum} for k, c in cnt.items()}


def section_of(r, code, is_dup=False, big_perf=()):
    """라벨 윗줄에 찍을 섹션명."""
    if is_dup:
        return "중복음반"
    g = canon_genre(r.get("genre"))
    if g != "클래식":
        return SECTION_NAME.get(g, g or "미분류")
    if "박스앨범" in str(r.get("notes") or ""):
        return "클래식 · 박스앨범"      # 부피가 커서 따로 세우는 묶음
    is_cg = re.sub(r"^(?:LP|CD)-", "", str(code or "")).startswith("CG")
    head = "그라모폰" if is_cg else "클래식"
    perf = primary(lead_artist(r.get("performer") or r.get("artist"), g))
    if isinstance(big_perf, dict):
        big_perf = big_perf.get("CG" if is_cg else "C", ())
    if perf in big_perf:
        return f"{head} · {perf}"
    comp = str(r.get("composer") or "").strip()
    words = name_words(comp)
    name = words[-1] if words else ""
    return f"{head} · {'모음집' if not name or name.upper() == 'VA' else name}"


def build(rows, albums, album_code, digital):
    dup_moves = {}
    for key, idxs in albums.items():
        lp = [i for i in idxs if rows[i].get("medium") == "LP"]
        if len(lp) < 2:
            continue
        in_main = [i for i in lp if rows[i]["sheet"] in MAIN]
        keep = in_main[0] if in_main else lp[0]
        for i in lp:
            dup_moves[i] = ("보관(대표본, 중복 %d매)" % len(lp)) if i == keep \
                else "이동→보조(우측/침대옆/박스)"

    wb = Workbook()
    wb.remove(wb.active)
    by_sheet = defaultdict(list)
    for i, r in enumerate(rows):
        by_sheet[r["sheet"]].append((i, r))
    for name in SHEET_ORDER:
        if name not in by_sheet:
            continue
        ws = wb.create_sheet(name[:31])
        head(ws, HDR, WID)
        for seq, (i, r) in enumerate(by_sheet[name], start=1):
            key = album_key(r)
            code = album_code.get(key, "")
            # 같은 앨범을 LP·CD 둘 다 소장하면 실물(행)의 매체로 접두를 맞춘다
            med = str(r.get("medium") or "").strip().upper()
            if code and med in ("LP", "CD"):
                code = re.sub(r"^(?:LP|CD)-", "", code)
                code = f"{med}-{code}"
            vals = []
            for k in KEYS:
                if k == "order":
                    vals.append(seq)
                elif k == "code":
                    vals.append(code)
                elif k == "digital":
                    vals.append("✓" if key in digital else "")
                elif k == "dup":
                    vals.append(dup_moves.get(i, ""))
                else:
                    vals.append(r.get(k, ""))
            ws.append(vals)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HDR))}{ws.max_row}"

    ws = wb.create_sheet("분류코드표")
    head(ws, ["분류코드", "장르", "아티스트/작곡가-연주자", "앨범", "매수", "디지털"],
         [14, 10, 34, 36, 6, 8])
    for key in sorted(album_code, key=lambda k: album_code[k]):
        idxs = albums[key]
        r = rows[idxs[0]]
        who = (f'{r.get("composer")} - {r.get("performer")}'
               if str(r.get("composer") or "").strip() else r.get("artist"))
        ws.append([album_code[key], r.get("genre", ""), who, r.get("album", ""),
                   len(idxs), "✓" if key in digital else ""])

    ws = wb.create_sheet("중복목록")
    head(ws, ["분류코드", "아티스트", "앨범", "매수", "위치들", "대표본 위치"],
         [14, 24, 32, 6, 40, 18])
    for key, idxs in sorted(albums.items(), key=lambda kv: -len(kv[1])):
        if len(idxs) < 2:
            continue
        r = rows[idxs[0]]
        keep = next((rows[i]["sheet"] for i in idxs if "보관" in dup_moves.get(i, "")),
                    rows[idxs[0]]["sheet"])
        ws.append([album_code.get(key, ""), r.get("artist", ""), r.get("album", ""),
                   len(idxs), ", ".join(rows[i]["sheet"] for i in idxs), keep])

    ws = wb.create_sheet("미식별목록")
    head(ws, ["시트(위치)", "엑셀행", "순번", "매체", "보유 정보", "단서", "앞 이웃", "뒤 이웃"],
         [16, 7, 6, 6, 34, 48, 26, 26])
    for name in SHEET_ORDER:
        lst = by_sheet.get(name, [])
        for pos, (i, r) in enumerate(lst):
            a, al = r.get("artist"), r.get("album")
            if "비매품" in str(al or "") + str(a or ""):
                continue
            if not blankv(a) and not blankv(al):
                continue
            have = (f"아티스트만: {a}" if not blankv(a)
                    else (f"앨범만: {al}" if not blankv(al) else "정보 없음"))
            clues = [x for x in (r.get("label_cat"), r.get("db_url"), r.get("notes"))
                     if not blankv(x)]

            def nb(rng):
                for j in rng:
                    if 0 <= j < len(lst):
                        a2, l2 = lst[j][1].get("artist"), lst[j][1].get("album")
                        if not blankv(a2) and not blankv(l2):
                            return f"{str(a2)[:12]}—{str(l2)[:16]}"
                return "(없음)"
            ws.append([name, pos + 2, pos + 1, r.get("medium", ""), have,
                       " | ".join(str(c)[:46] for c in clues[:2]) or "(단서 없음)",
                       nb(range(pos - 1, -1, -1)), nb(range(pos + 1, len(lst)))])

    lp_genre = defaultdict(int)
    for r in rows:
        if r.get("medium") == "LP" and str(r.get("artist") or "").strip():
            lp_genre[str(r.get("genre") or "").strip() or "미상"] += 1
    ws = wb.create_sheet("정리계획")
    head(ws, ["장르(LP)", "매수", "배치 제안(메인 정리장 순서대로)"], [14, 8, 90])
    shelves = ["LP좌측선반1층", "LP좌측선반2층", "LP좌측선반3층",
               "LP중앙선반1열1층", "LP중앙선반1열2층", "LP중앙선반1열3층",
               "LP중앙선반2열1층", "LP중앙선반2열2층", "LP중앙선반2열3층"]
    si, remaining = 0, CAPACITY[shelves[0]]
    for g, cnt in sorted(lp_genre.items(), key=lambda x: -x[1]):
        need, spots = cnt, []
        while need > 0 and si < len(shelves):
            take = min(need, remaining)
            spots.append(f"{shelves[si]}({take})")
            need -= take
            remaining -= take
            if remaining == 0:
                si += 1
                if si < len(shelves):
                    remaining = CAPACITY[shelves[si]]
        ws.append([g, cnt, " + ".join(spots) + ("  ⚠잔여 %d→보조" % need if need else "")])
    ws.append([])
    ws.append(["※ 용량=현재 적재량 기준(중앙2열1층만 80). 좌측선반3층=DG 2864(CG) 전용. "
               "중복 대표본만 메인, 나머지는 보조로."])

    # 라벨인쇄: 매체별 시트 분리(LP 먼저 작업 예정) — 열당 136행, 위치 순서
    ROWS = 136
    label_sets = {"라벨인쇄-LP": [], "라벨인쇄-CD": []}
    for name in SHEET_ORDER:
        for i, r in by_sheet.get(name, []):
            c = album_code.get(album_key(r), "")
            if not c:
                continue
            med = str(r.get("medium") or "").strip().upper()
            tab = "라벨인쇄-LP" if med == "LP" else ("라벨인쇄-CD" if med == "CD" else None)
            if tab:
                label_sets[tab].append(f"{med}-" + re.sub(r"^(?:LP|CD)-", "", c))
    label_sets["라벨인쇄-전체"] = (label_sets["라벨인쇄-LP"]
                                 + label_sets["라벨인쇄-CD"])   # LP 먼저, 이어서 CD
    for tab, codes in label_sets.items():
        ws = wb.create_sheet(tab)
        for idx, code in enumerate(codes):
            cell = ws.cell(row=idx % ROWS + 1, column=idx // ROWS + 1, value=code)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(1, max(1, (len(codes) + ROWS - 1) // ROWS) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 16

    # 파이프라인이 만들지 않는 사용자 시트(정리작업·라벨재발급N 등)는 그대로 옮겨온다.
    # (워크북을 새로 만들기 때문에 명시적으로 복사하지 않으면 사라진다)
    generated = set(SHEET_ORDER) | {"분류코드표", "중복목록", "미식별목록", "정리계획",
                                    "라벨인쇄", "라벨인쇄-LP", "라벨인쇄-CD", "라벨인쇄-전체"}
    if os.path.exists(XLSX):
        old = openpyxl.load_workbook(XLSX)
        for name in old.sheetnames:
            if name in generated:
                continue
            src, dst = old[name], wb.create_sheet(name)
            for row in src.iter_rows():
                for cell in row:
                    nc = dst.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        nc.font, nc.fill = copy(cell.font), copy(cell.fill)
                        nc.alignment = copy(cell.alignment)
            for k, dim in src.column_dimensions.items():
                dst.column_dimensions[k].width = dim.width
            dst.freeze_panes = src.freeze_panes

    for w in wb.worksheets:
        if w.title.startswith("라벨인쇄"):
            continue
        hdr = [c.value for c in w[1]]
        for ci, h in enumerate(hdr, start=1):
            horiz = "center" if h in CENTER else "left"
            for (cell,) in w.iter_rows(min_row=2, min_col=ci, max_col=ci):
                cell.alignment = Alignment(horizontal=horiz, vertical="top", wrap_text=True)
    return wb, (len(label_sets["라벨인쇄-LP"]), len(label_sets["라벨인쇄-CD"]))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="코드 미리보기만, 파일 미변경")
    ap.add_argument("--export-only", action="store_true",
                    help="엑셀·라벨 재작성 없이 data/ 산출물(code_registry·physical_albums)만 생성. "
                         "새 기계 부트스트랩용 — 엑셀 정본을 건드리지 않는다")
    args = ap.parse_args()

    rows = extract(XLSX)
    reg = load_registry()
    first_seed = not reg["artists"] and not reg["composers"]
    if first_seed:
        seed_registry_from_existing(rows, reg)
        print("· 레지스트리 최초 이식(기존 코드 고정) 완료")
    known = {c for g in reg["albums"].values() for c in g} | \
            {c for g in reg["perf_albums"].values() for c in g}

    warnings = []
    albums, album_code, meta = assign_codes(rows, reg, warnings)

    new_codes = []
    for key, code in album_code.items():
        if key[1] not in known or first_seed and False:
            pass
    # 신규 = 이번에 레지스트리에 처음 들어간 앨범
    known_after = {c for g in reg["albums"].values() for c in g} | \
                  {c for g in reg["perf_albums"].values() for c in g}
    fresh = known_after - known if not first_seed else set()
    for key, code in sorted(album_code.items(), key=lambda kv: kv[1]):
        if key[1] in fresh:
            r = rows[albums[key][0]]
            new_codes.append((code, r.get("artist"), r.get("album"), r["sheet"]))

    digital = {}
    if os.path.exists(DIGITAL):
        digital = {tuple(k.split("|", 1)): v
                   for k, v in json.load(open(DIGITAL, encoding="utf-8")).items()}

    print(f"· 행 {len(rows)} | 코드 {len(album_code)}종 | 신규 {len(new_codes)}건"
          f" | 디지털 연동 {len(digital)}건")
    for w in warnings:
        print("  ⚠", w)
    if new_codes:
        print("· 신규 분류코드 (라벨 추가 인쇄 대상):")
        for code, a, al, sh in new_codes:
            print(f"    {code:<16} {str(a)[:20]:<20} {str(al)[:28]:<28} @{sh}")

    if args.dry_run:
        print("· dry-run — 파일 미변경")
        return
    nlabels = None
    if not args.export_only:
        wb, nlabels = build(rows, albums, album_code, digital)
        wb.save(XLSX)
    # 실물 위키(md) 생성용 구조화 데이터 내보내기
    out = []
    for key, code in album_code.items():
        idxs = albums[key]
        r = rows[idxs[0]]
        out.append({
            "code": code, "genre": r.get("genre", ""), "artist": r.get("artist", ""),
            "album": r.get("album", ""), "rep": r.get("rep", ""),
            "composer": r.get("composer", ""), "performer": r.get("performer", ""),
            "desc": r.get("desc", ""), "discogs_id": str(r.get("discogs_id", "")),
            "db_url": r.get("db_url", ""), "label_cat": r.get("label_cat", ""),
            "media": sorted({rows[j].get("medium", "") for j in idxs if rows[j].get("medium")}),
            "locations": sorted({rows[j]["sheet"] for j in idxs}),
            "copies": len(idxs), "digital": key in digital,
        })
    pj = os.path.join(INV, "data", "physical_albums.json")
    os.makedirs(os.path.dirname(pj), exist_ok=True)
    json.dump(sorted(out, key=lambda x: x["code"]),
              open(pj, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"· physical_albums.json {len(out)}건 내보냄")
    json.dump(reg, open(REGISTRY, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    if nlabels is None:
        print("· 레지스트리 갱신 (--export-only: 엑셀·라벨 미변경)")
    else:
        print(f"· 저장: {XLSX} (라벨 LP {nlabels[0]} / CD {nlabels[1]} / 전체 {nlabels[0] + nlabels[1]}) / 레지스트리 갱신")


if __name__ == "__main__":
    main()
