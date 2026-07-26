#!/usr/bin/env python3
"""v3 workbook build from lpcd_v3.json (extracted from the user-edited xlsx).

Changes vs v2: leading-"The" ignored for filing (The Beatles → B); DG yellow
2864-series gets prefix CG (kept apart, mostly 좌측선반3층); 순번 renumbered
1..n per sheet; per-column horizontal alignment made consistent.
"""
import json
import os
import re
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCR = ("/tmp/claude-1000/-home-neotango-media-archive-music-wiki/"
       "4e6dd816-7bba-408a-a949-f563a3925c6f/scratchpad")
OUT = "/home/neotango/lpcd_image/20260726_LPCD_목록_v2.xlsx"

GENRE_LETTER = {"클래식": "C", "클래식기타": "G", "재즈": "J", "탱고": "T",
                "월드": "W", "팝": "P", "가요": "K", "OST·경음악": "O"}
SHEET_ORDER = ["CD윗층", "CD아래층",
               "LP중앙선반1열1층", "LP중앙선반1열2층", "LP중앙선반1열3층",
               "LP중앙선반2열1층", "LP중앙선반2열2층", "LP중앙선반2열3층",
               "LP좌측선반1층", "LP좌측선반2층", "LP좌측선반3층",
               "LP우측선반1층", "LP우측선반2층", "LP보관박스1", "LP보관박스2", "LP침대옆"]
MAIN = {"LP중앙선반1열1층", "LP중앙선반1열2층", "LP중앙선반1열3층", "LP중앙선반2열1층",
        "LP중앙선반2열2층", "LP중앙선반2열3층", "LP좌측선반1층", "LP좌측선반2층", "LP좌측선반3층"}
CAPACITY = {"LP중앙선반1열1층": 79, "LP중앙선반1열2층": 55, "LP중앙선반1열3층": 59,
            "LP중앙선반2열1층": 80, "LP중앙선반2열2층": 58, "LP중앙선반2열3층": 86,
            "LP좌측선반1층": 46, "LP좌측선반2층": 61, "LP좌측선반3층": 65}

CHO = "ㄱㄲㄴㄷㄸㄹㅁㅂㅃㅅㅆㅇㅈㅉㅊㅋㅌㅍㅎ"
CHO_BASE = {"ㄲ": "ㄱ", "ㄸ": "ㄷ", "ㅃ": "ㅂ", "ㅆ": "ㅅ", "ㅉ": "ㅈ"}


def initial(name: str) -> str:
    for ch in str(name or "").strip():
        if "가" <= ch <= "힣":
            c = CHO[(ord(ch) - 0xAC00) // 588]
            return CHO_BASE.get(c, c)
        if ch.isascii() and ch.isalpha():
            return ch.upper()
        if ch.isdigit():
            return "0"
    return "X"


def artist_key(name: str) -> str:
    """Filing form of an artist name: leading articles ignored.
    The Beatles→Beatles, El Quinteto→Quinteto, Los Romeros→Romeros.
    Exceptions kept: 'Los Angeles …'(지명), 'I Musici'류(관례상 이름 일부)."""
    s = str(name or "")
    s = re.sub(r"^\s*the\s+", "", s, flags=re.I)
    s = re.sub(r"^\s*el\s+", "", s, flags=re.I)
    if re.match(r"^\s*los\s+(?!angeles\b)", s, flags=re.I):
        s = re.sub(r"^\s*los\s+", "", s, flags=re.I)
    return s


def norm(s):
    return re.sub(r"[^a-z0-9가-힣]", "", str(s or "").lower())


def surname_key(composer: str) -> str:
    words = re.findall(r"[A-Za-z가-힣]+", str(composer or ""))
    return words[-1].lower() if words else norm(composer)


def surname_initial(name: str) -> str:
    s = str(name or "").strip()
    if any("가" <= ch <= "힣" for ch in s):
        return initial(s)
    words = [w for w in re.findall(r"[A-Za-z]+", s) if w.lower() != "von"]
    return words[-1][0].upper() if words else initial(s)


rows = json.load(open(os.path.join(SCR, "lpcd_v3.json"), encoding="utf-8"))

albums = defaultdict(list)
for i, r in enumerate(rows):
    a, al = str(r.get("artist") or "").strip(), str(r.get("album") or "").strip()
    if a and al:
        albums[(norm(a), norm(al))].append(i)

art_group = defaultdict(set)
art_albums = defaultdict(set)
art_display = {}
comp_group = defaultdict(set)        # (pref, init) -> ck
comp_display = {}
perf_group = defaultdict(set)        # (pref, ck, init) -> pk
perf_display = {}
perf_albums = defaultdict(set)

for key, idxs in albums.items():
    r = rows[idxs[0]]
    g = str(r.get("genre") or "").strip()
    letter = GENRE_LETTER.get(g, "X")
    if letter == "C" and (r.get("composer") or r.get("performer")):
        # DG yellow gatefold 2864-series → CG (kept physically separate)
        pref = "CG" if any("2864" in str(rows[j].get("label_cat") or "")
                           or (rows[j]["sheet"] == "LP좌측선반3층")
                           for j in idxs) else "C"
        comp = str(r.get("composer") or "VA").strip() or "VA"
        perf = str(r.get("performer") or r.get("artist") or "").strip()
        ck = (pref, surname_key(comp) or "va")
        comp_group[(pref, surname_initial(comp) if comp != "VA" else "V")].add(ck)
        comp_display.setdefault(ck, comp)
        pk = (ck, norm(perf) or "unknown")
        perf_group[(ck, surname_initial(perf))].add(pk)
        perf_display.setdefault(pk, perf)
        perf_albums[pk].add(key)
        r["_cls"] = ("C", pref, ck, pk)
    else:
        # 한국 비매품(프로모션) 음반은 K → KN 으로 분리
        if letter == "K" and any(
                "비매품" in (str(rows[j].get("album") or "") + str(rows[j].get("artist") or "")
                           + str(rows[j].get("notes") or "")) for j in idxs):
            letter = "KN"
        ak = (letter, norm(artist_key(r.get("artist"))))
        art_group[(letter, initial(artist_key(r.get("artist"))))].add(ak)
        art_display.setdefault(ak, r.get("artist"))
        art_albums[ak].add(key)
        r["_cls"] = ("N", ak)

art_no = {}
for (letter, init), aset in art_group.items():
    gap = 3 if len(aset) > 20 else 4
    for n, ak in enumerate(sorted(aset, key=lambda k: norm(artist_key(art_display[k])))):
        art_no[ak] = 1 + gap * n
comp_no = {}
for (pref, init), cset in comp_group.items():
    for n, ck in enumerate(sorted(cset)):
        comp_no[ck] = n
perf_no = {}
for (ck, init), pset in perf_group.items():
    for n, pk in enumerate(sorted(pset, key=lambda k: norm(perf_display[k]))):
        perf_no[pk] = 1 + 4 * n

album_code = {}
for key, idxs in albums.items():
    r = rows[idxs[0]]
    cls = r.get("_cls")
    if not cls:
        continue
    if cls[0] == "N":
        _, ak = cls
        keys = sorted(art_albums[ak], key=lambda k: k[1])
        an = keys.index(key) + 1
        album_code[key] = (f"{ak[0]}-{initial(artist_key(art_display[ak]))}"
                           f"{art_no[ak]:02d}-{an:02d}")
    else:
        _, pref, ck, pk = cls
        comp = comp_display[ck]
        cinit = "V" if comp == "VA" else surname_initial(comp)
        keys = sorted(perf_albums[pk], key=lambda k: k[1])
        an = keys.index(key) + 1
        album_code[key] = (f"{pref}-{cinit}{comp_no[ck]}-"
                           f"{surname_initial(perf_display[pk])}{perf_no[pk]:02d}-{an:02d}")

dup_moves = {}
for key, idxs in albums.items():
    lp = [i for i in idxs if rows[i].get("medium") == "LP"]
    if len(lp) < 2:
        continue
    in_main = [i for i in lp if rows[i]["sheet"] in MAIN]
    keep = in_main[0] if in_main else lp[0]
    for i in lp:
        if i != keep:
            dup_moves[i] = "이동→보조(우측/침대옆/박스)"
    dup_moves[keep] = f"보관(대표본, 중복 {len(lp)}매)"

lp_genre = defaultdict(int)
for r in rows:
    if r.get("medium") == "LP" and str(r.get("artist") or "").strip():
        lp_genre[str(r.get("genre") or "").strip() or "미상"] += 1
main_shelves = ["LP좌측선반1층", "LP좌측선반2층", "LP좌측선반3층",
                "LP중앙선반1열1층", "LP중앙선반1열2층", "LP중앙선반1열3층",
                "LP중앙선반2열1층", "LP중앙선반2열2층", "LP중앙선반2열3층"]
zone_rows = []
si = 0
remaining = CAPACITY[main_shelves[0]]
for g, cnt in sorted(lp_genre.items(), key=lambda x: -x[1]):
    need = cnt
    spots = []
    while need > 0 and si < len(main_shelves):
        take = min(need, remaining)
        spots.append(f"{main_shelves[si]}({take})")
        need -= take
        remaining -= take
        if remaining == 0:
            si += 1
            if si < len(main_shelves):
                remaining = CAPACITY[main_shelves[si]]
    zone_rows.append((g, cnt, " + ".join(spots) + ("  ⚠잔여 %d→보조" % need if need > 0 else "")))

HDR = ["순번", "분류코드", "매체", "장르", "아티스트", "앨범제목", "대표곡",
       "작곡가", "연주자", "레이블/카탈로그번호", "Discogs ID", "MusicBrainz MBID",
       "DB링크", "중복/이동", "해설", "비고"]
KEYS = ["order", "code", "medium", "genre", "artist", "album", "rep",
        "composer", "performer", "label_cat", "discogs_id", "mbid",
        "db_url", "dup", "desc", "notes"]
WID = [6, 14, 6, 10, 24, 32, 26, 14, 16, 20, 12, 26, 38, 22, 44, 22]
CENTER = {"순번", "분류코드", "매체", "장르", "Discogs ID", "매수", "엑셀행"}
HF, HFo = PatternFill("solid", fgColor="305496"), Font(bold=True, color="FFFFFF")

wb = Workbook()
wb.remove(wb.active)


def head(ws, headers, widths=None):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HF, HFo
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if widths:
            ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    ws.freeze_panes = "A2"


by_sheet = defaultdict(list)
for i, r in enumerate(rows):
    by_sheet[r["sheet"]].append((i, r))
for name in SHEET_ORDER:
    if name not in by_sheet:
        continue
    ws = wb.create_sheet(name[:31])
    head(ws, HDR, WID)
    for seq, (i, r) in enumerate(by_sheet[name], start=1):
        a, al = norm(r.get("artist")), norm(r.get("album"))
        code = album_code.get((a, al), "") if a and al else ""
        vals = []
        for k in KEYS:
            if k == "order":
                vals.append(seq)          # 순번 재부여(현재 행 순서 기준)
            elif k == "code":
                vals.append(code)
            elif k == "dup":
                vals.append(dup_moves.get(i, ""))
            else:
                vals.append(r.get(k, ""))
        ws.append(vals)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HDR))}{ws.max_row}"

ws = wb.create_sheet("분류코드표")
head(ws, ["분류코드", "장르", "아티스트/작곡가-연주자", "앨범", "매수"], [14, 10, 34, 36, 6])
for key in sorted(album_code, key=lambda k: album_code[k]):
    idxs = albums[key]
    r = rows[idxs[0]]
    who = (f'{r.get("composer")} - {r.get("performer")}'
           if r.get("_cls", ("N",))[0] == "C" else r.get("artist"))
    ws.append([album_code[key], r.get("genre", ""), who, r.get("album", ""), len(idxs)])

ws = wb.create_sheet("중복목록")
head(ws, ["분류코드", "아티스트", "앨범", "매수", "위치들", "대표본 위치"], [14, 24, 32, 6, 40, 18])
for key, idxs in sorted(albums.items(), key=lambda kv: -len(kv[1])):
    if len(idxs) < 2:
        continue
    r = rows[idxs[0]]
    keep = next((rows[i]["sheet"] for i in idxs if "보관" in dup_moves.get(i, "")),
                rows[idxs[0]]["sheet"])
    ws.append([album_code.get(key, ""), r.get("artist", ""), r.get("album", ""),
               len(idxs), ", ".join(rows[i]["sheet"] for i in idxs), keep])

ws = wb.create_sheet("미식별목록")
head(ws, ["시트(위치)", "엑셀행", "순번", "매체", "보유 정보", "단서", "앞 이웃", "뒤 이웃"],
     [16, 7, 6, 6, 34, 48, 26, 26])


def blankv(v):
    return not str(v or "").strip()


for name in SHEET_ORDER:
    if name not in by_sheet:
        continue
    lst = by_sheet[name]
    for pos, (i, r) in enumerate(lst):
        a, al = r.get("artist"), r.get("album")
        if "비매품" in str(al or "") + str(a or ""):
            continue
        if not blankv(a) and not blankv(al):
            continue
        have = (f"아티스트만: {a}" if not blankv(a)
                else (f"앨범만: {al}" if not blankv(al) else "정보 없음"))
        clues = [x for x in (r.get("label_cat"), r.get("db_url"), r.get("notes")) if not blankv(x)]
        clue = " | ".join(str(c)[:46] for c in clues[:2]) or "(단서 없음)"

        def nb(rng):
            for j in rng:
                if 0 <= j < len(lst):
                    a2, l2 = lst[j][1].get("artist"), lst[j][1].get("album")
                    if not blankv(a2) and not blankv(l2):
                        return f"{str(a2)[:12]}—{str(l2)[:16]}"
            return "(없음)"
        ws.append([name, pos + 2, pos + 1, r.get("medium", ""), have, clue,
                   nb(range(pos - 1, -1, -1)), nb(range(pos + 1, len(lst)))])

ws = wb.create_sheet("정리계획")
head(ws, ["장르(LP)", "매수", "배치 제안(메인 정리장 순서대로)"], [14, 8, 90])
for g, cnt, plan in zone_rows:
    ws.append([g, cnt, plan])
ws.append([])
ws.append(["※ 용량=현재 적재량 기준(중앙2열1층만 80). 좌측선반3층=DG 2864 시리즈(CG) 전용 유지. "
           "중복 대표본만 메인, 나머지는 보조(우측/침대옆/박스1)로."])

# 열별 일관 정렬(가운데: 짧은 코드/번호, 좌측: 텍스트)
for w in wb.worksheets:
    hdr = [c.value for c in w[1]]
    for ci, h in enumerate(hdr, start=1):
        horiz = "center" if h in CENTER else "left"
        for (cell,) in w.iter_rows(min_row=2, min_col=ci, max_col=ci):
            cell.alignment = Alignment(horizontal=horiz, vertical="top", wrap_text=True)

# 라벨인쇄: 위치 순서대로 코드만, 열당 136행 (라벨 프로그램 형식)
ws = wb.create_sheet("라벨인쇄")
codes = []
for name in SHEET_ORDER:
    for i, r in by_sheet.get(name, []):
        a, al = norm(r.get("artist")), norm(r.get("album"))
        c = album_code.get((a, al), "") if a and al else ""
        if c:
            codes.append(c)
ROWS = 136
for idx, code in enumerate(codes):
    cell = ws.cell(row=idx % ROWS + 1, column=idx // ROWS + 1, value=code)
    cell.alignment = Alignment(horizontal="center", vertical="center")
for c in range(1, (len(codes) + ROWS - 1) // ROWS + 1):
    ws.column_dimensions[get_column_letter(c)].width = 13

wb.save(OUT)
cg = sum(1 for v in album_code.values() if v.startswith("CG-"))
print(f"wrote {OUT}")
print(f"  분류코드 {len(album_code)}종 (CG {cg}종) | 중복이동 {sum(1 for v in dup_moves.values() if v.startswith('이동'))}매 | 시트 {len(wb.sheetnames)}")
