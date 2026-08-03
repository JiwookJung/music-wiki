#!/usr/bin/env python3
"""선반 배치·섹션표·라벨 재출력 목록 생성 (pipeline.py 실행 뒤에 돌린다).

  python build_layout.py <최초_라벨출력분.xlsx>

배치-* 13장  : 선반별로 어떤 음반을 어느 순서로 꽂을지. 섹션 → 분류코드 순.
섹션표       : 구분판에 붙일 섹션명과 코드 범위.
라벨재출력    : 최초 라벨 출력분과 코드가 달라진 것 전량(섹션 포함).

섹션 계산은 pipeline.section_of 하나만 쓴다. 라벨 시트(pipeline)와 배치 시트가
서로 다른 함수를 쓰면 같은 음반에 다른 섹션이 찍힌다.
"""

import importlib.util
import re
import sys

import openpyxl
from collections import Counter, defaultdict
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

S = sys.argv[1]
spec = importlib.util.spec_from_file_location("p", "inventory/scripts/pipeline.py")
p = importlib.util.module_from_spec(spec)
spec.loader.exec_module(p)
V3 = "/home/neotango/lpcd_image/20260801_LPCD_목록_v3.xlsx"
def norm(s):
    return re.sub(r"[^0-9a-z가-힣]", "", str(s or "").lower())
wb = openpyxl.load_workbook(V3)
COLS = [
    "분류코드",
    "장르",
    "아티스트",
    "앨범제목",
    "대표곡",
    "작곡가",
    "연주자",
    "레이블/카탈로그번호",
    "비고",
    "해설",
]
reps, dups = [], []
for name in wb.sheetnames:
    if not p.is_location_sheet(name) or not name.startswith("LP") or name.startswith("라벨"):
        continue
    ws = wb[name]
    hdr = [str(x.value or "") for x in ws[1]]
    if "분류코드" not in hdr:
        continue
    idx = {c: hdr.index(c) for c in COLS if c in hdr}
    idp = hdr.index("중복/이동")
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[idx["분류코드"]]:
            continue
        rec = {c: (str(r[i]) if r[i] is not None else "") for c, i in idx.items()}
        rec["현재위치"] = name.replace("LP", "")
        rec.update(
            genre=rec["장르"],
            composer=rec.get("작곡가", ""),
            performer=rec.get("연주자", ""),
            artist=rec.get("아티스트", ""),
            notes=rec.get("비고", ""),
        )
        (dups if "이동" in str(r[idp] or "") else reps).append(rec)
cnt = {"C": Counter(), "CG": Counter()}
for r in reps:
    if p.canon_genre(r["장르"]) != "클래식":
        continue
    pref = "CG" if re.sub(r"^(?:LP|CD)-", "", r["분류코드"]).startswith("CG") else "C"
    pn = p.primary(p.lead_artist(r["performer"] or r["artist"], "클래식"))
    if norm(pn) and norm(pn) not in p._NOT_PERF:
        cnt[pref][pn] += 1
BIG = {k: {n for n, v in c.items() if v >= p.BIG_PERF_MIN} for k, c in cnt.items()}
def sec(r, d):
    """섹션 계산은 파이프라인 함수 하나만 쓴다(라벨 시트와 어긋나면 안 된다)."""
    return p.section_of(r, r["분류코드"], d, BIG)

def by(rs):
    return sorted(rs, key=lambda x: x["분류코드"])
def g(name):
    return by([r for r in reps if r["장르"] == name])
클래식 = g("클래식")
CG = [r for r in 클래식 if r["분류코드"].startswith("LP-CG")]
보관 = [r for r in 클래식 if "Heritage of Music" in r["앨범제목"]]
일반C = [r for r in 클래식 if r not in CG and r not in 보관]
BOX = re.compile(r"[Bb]ox|박스|전집|\d\s*LPs?\b|Complete")
박스 = [r for r in 일반C if "박스앨범" in r.get("비고", "")]
비박스 = [r for r in 일반C if r not in 박스]
가요 = g("가요")
재즈 = g("재즈")
OST = g("OST·경음악")
기타 = g("클래식기타")
경음악 = g("경음악")
팝 = g("팝")
탱고 = g("탱고")
월드 = g("월드")
가요중복 = by([r for r in dups if r["장르"] == "가요"])
비가요중복 = by([r for r in dups if r["장르"] != "가요"])
# 클래식(비박스)은 섹션이 선반 경계에서 잘리지 않도록 '섹션 통째로' 채운다.
grp = defaultdict(list)
for r in 비박스:
    grp[sec(r, False)].append(r)
secs = sorted(grp, key=lambda k: min(x["분류코드"] for x in grp[k]))
CAPS = [("침대옆", 52), ("우측1층", 45 - len(박스)), ("우측2층", 18)]
# 섹션은 통째로 옮기되, 잡탕인 '모음집' 만은 선반 경계에서 쪼갠다.
pack = {n: [] for n, _ in CAPS}
queue = [(k, sorted(grp[k], key=lambda x: x["분류코드"])) for k in secs]
qi = 0
for name, cap in CAPS:
    while qi < len(queue):
        k, items = queue[qi]
        room = cap - len(pack[name])
        if len(items) <= room:
            pack[name] += items
            qi += 1
            continue
        if k.endswith("모음집") and room > 0:  # 잡탕만 분할 허용
            pack[name] += items[:room]
            queue[qi] = (k, items[room:])
        break
for k, items in queue[qi:]:
    pack[CAPS[-1][0]] += items  # 남으면 마지막 선반
j3 = 21
k3, k2 = 63, 59
PLAN = [
    ("배치-좌측3층", 65, CG, []),
    ("배치-좌측2층", 61, 팝, 비가요중복),
    ("배치-좌측1층", 53, 기타, []),
    ("배치-중앙좌3층", 63, 가요[:k3], []),
    ("배치-중앙좌2층", 59, 가요[k3 : k3 + k2], []),
    ("배치-중앙좌1층", 84, 가요[k3 + k2 :], []),
    ("배치-중앙우3층", 86, 탱고 + 월드 + 재즈[:j3], []),
    ("배치-중앙우2층", 58, 재즈[j3:], []),
    ("배치-중앙우1층", 80, OST + 경음악, []),
    ("배치-침대옆", 52, pack["침대옆"], []),
    ("배치-우측1층", 45, 박스 + pack["우측1층"], []),
    ("배치-우측2층", 18, pack["우측2층"], []),
    ("배치-보관함", 67, 보관, 가요중복),
]
HDR = [
    "순번",
    "섹션",
    "분류코드",
    "장르",
    "아티스트",
    "앨범제목",
    "대표곡",
    "작곡가",
    "연주자",
    "레이블/카탈로그번호",
    "현재위치",
    "이동필요",
    "비고",
    "해설",
]
mv = 0
secmap = {}
shelfsec = []
for nm, cap, normal, dup in PLAN:
    if nm in wb.sheetnames:
        del wb[nm]
    ws = wb.create_sheet(nm)
    ws.append(HDR)
    for c in range(1, len(HDR) + 1):
        cl = ws.cell(row=1, column=c)
        cl.fill = PatternFill("solid", fgColor="305496")
        cl.font = Font(bold=True, color="FFFFFF")
    dest = nm.replace("배치-", "").replace("선반", "")
    order = defaultdict(list)
    for r, d in [(x, False) for x in normal] + [(x, True) for x in dup]:
        order[sec(r, d)].append(r)
    def keyf(s):
        return (s == "중복음반", min(x["분류코드"] for x in order[s]))
    seq = [
        (s, r)
        for s in sorted(order, key=keyf)
        for r in sorted(order[s], key=lambda x: x["분류코드"])
    ]
    for i, (s, r) in enumerate(seq, 1):
        cur = (
            r["현재위치"]
            .replace("선반", "")
            .replace("중앙1열", "중앙좌")
            .replace("중앙2열", "중앙우")
        )
        cur = re.sub(r"^보관박스\d$", "보관함", cur)
        moved = "" if cur == dest else "이동"
        mv += 1 if moved else 0
        secmap[r["분류코드"]] = (s, nm.replace("배치-", ""), i)
        ws.append(
            [
                i,
                s,
                r["분류코드"],
                r["장르"],
                r["아티스트"],
                r["앨범제목"],
                r.get("대표곡", ""),
                r.get("작곡가", ""),
                r.get("연주자", ""),
                r.get("레이블/카탈로그번호", ""),
                r["현재위치"],
                moved,
                r.get("비고", ""),
                r.get("해설", ""),
            ]
        )
    for s in sorted(order, key=keyf):
        cs = sorted(x["분류코드"] for x in order[s])
        shelfsec.append(
            (nm.replace("배치-", ""), cap, s, len(cs), "○" if len(cs) >= 2 else "", cs[0], cs[-1])
        )
    for w, c in zip([6, 26, 16, 10, 22, 30, 24, 18, 20, 20, 14, 8, 26, 60], range(1, 15)):
        ws.column_dimensions[get_column_letter(c)].width = w
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    print(f"  {nm.replace('배치-', ''):<10}{cap:>4}칸 {len(seq):>4}장 섹션 {len(order):>3}개")


def mk(title, rows, hdr, widths, notes, color, idx):
    if title in wb.sheetnames:
        del wb[title]
    w = wb.create_sheet(title, idx)
    w.append([notes[0]])
    w.append(hdr)
    for r in rows:
        w.append(list(r))
    w.append([])
    for t in notes[1:]:
        w.append([t])
    for c in range(1, len(hdr) + 1):
        w.cell(row=2, column=c).fill = PatternFill("solid", fgColor=color)
        w.cell(row=2, column=c).font = Font(bold=True, color="FFFFFF")
    for ww, i in zip(widths, range(1, len(widths) + 1), strict=False):
        w.column_dimensions[get_column_letter(i)].width = ww
    for row in w.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    w.freeze_panes = "A3"


# ── 선반에 붙일 구분판 라벨 ─────────────────────────────────────────────
# 음반 라벨과 달리 선반 라벨은 손으로 세우는 것이라 20장 안쪽이어야 쓸모가 있다.
# 규칙: 선반마다 최소 1장. 그 안에서 10장 이상인 섹션은 따로 1장 더.
DIVIDER_MIN = 10
shelf_secs = defaultdict(list)
for shelf, cap, name, n, mark, lo, hi in shelfsec:
    shelf_secs[shelf].append((name, n, lo, hi))
SHELF_TITLE = {  # 선반 전체를 가리키는 문구
    "좌측3층": "클래식 · 그라모폰 (DG 2864)",
    "중앙좌3층": "가요 ①", "중앙좌2층": "가요 ②", "중앙좌1층": "가요 ③",
    "침대옆": "클래식 ① (바흐부터)", "우측1층": "클래식 ②", "우측2층": "클래식 ③ (비발디까지)",
    "보관함": "중복 가요",
}
divs = []
for shelf, secs in shelf_secs.items():
    total = sum(x[1] for x in secs)
    big = [x for x in secs if x[1] >= DIVIDER_MIN]
    covered = sum(x[1] for x in big)
    lo, hi = min(x[2] for x in secs), max(x[3] for x in secs)
    # 큰 섹션 하나가 선반을 거의 다 채우면 선반 라벨 한 장으로 충분하다.
    if len(big) <= 1 and covered >= total * 0.9:
        title = SHELF_TITLE.get(shelf) or (big[0][0] if big else secs[0][0])
        divs.append((shelf, title, total, lo, hi, "선반 전체"))
        continue
    if covered < total:
        divs.append((shelf, SHELF_TITLE.get(shelf) or secs[0][0],
                     total, lo, hi, "선반 전체"))
    for name, n, slo, shi in sorted(big, key=lambda x: -x[1]):
        divs.append((shelf, name, n, slo, shi, "구분판"))
order = {n: i for i, n in enumerate(
    ["좌측1층", "좌측2층", "좌측3층", "중앙좌3층", "중앙좌2층", "중앙좌1층",
     "중앙우3층", "중앙우2층", "중앙우1층", "침대옆", "우측1층", "우측2층", "보관함"])}
divs.sort(key=lambda x: (order.get(x[0], 99), -x[2]))
divs = [(i,) + d for i, d in enumerate(divs, 1)]
mk(
    "선반라벨",
    divs,
    ["순번", "선반", "라벨 문구 (구분판에 인쇄)", "장수", "시작 코드", "끝 코드", "종류"],
    [5, 12, 34, 6, 18, 18, 10],
    [f"선반 구분판 라벨 — 총 {len(divs)}장",
     "※ 음반에 붙이는 라벨(라벨인쇄-LP)은 분류코드만 찍는다. 이 시트는 선반에 세울 구분판이다.",
     "※ '선반 전체'는 그 선반 맨 앞에, '구분판'은 해당 섹션이 시작하는 자리에 세운다.",
     f"※ {DIVIDER_MIN}장 미만인 작은 섹션은 구분판을 만들지 않는다(섹션표에서 코드 범위로 찾는다).",
     "※ 전체 섹션 목록은 '섹션표' 시트 참고."],
    "1F6F42",
    2,
)
mk(
    "섹션표",
    shelfsec,
    ["선반", "칸", "섹션", "장수", "구분판", "시작 코드", "끝 코드"],
    [13, 6, 34, 7, 7, 18, 18],
    [
        f"선반 섹션표 (2026-08-02) — 구분판에 붙일 이름 · 총 {len(shelfsec)}섹션 {sum(r[3] for r in shelfsec)}장",
        "※ 라벨 윗줄에 같은 섹션명이 찍혀 있다. 섹션 단위로 구분판을 세우면 된다.",
        "※ 클래식은 작곡가 섹션이 기본이고, 3장 이상 가진 연주자는 별도 섹션으로 뺐다.",
        "※ 중복음반은 장르와 무관하게 한 섹션으로 묶어 선반 끝에 둔다.",
        "※ 구분판 열이 비어 있으면 그 섹션은 1장뿐이라 구분판을 세울 필요가 없다.",
        "※ 연주자 섹션이 생기면 그 선반은 분류코드 오름차순이 아니라 섹션 순서로 꽂힌다.",
    ],
    "305496",
    0,
)


def load(path):
    w = openpyxl.load_workbook(path, read_only=True)
    out = []
    for n in w.sheetnames:
        if not p.is_location_sheet(n) or not n.startswith("LP") or n.startswith("라벨"):
            continue
        s = w[n]
        hdr = [str(x.value or "") for x in next(s.iter_rows(max_row=1))]
        if not {"분류코드", "아티스트", "앨범제목"} <= set(hdr):
            continue
        ix = {h: hdr.index(h) for h in ("분류코드", "아티스트", "앨범제목", "장르")}
        for r in s.iter_rows(min_row=2, values_only=True):
            if r and r[ix["분류코드"]]:
                out.append(
                    {
                        "code": str(r[ix["분류코드"]]),
                        "artist": str(r[ix["아티스트"]] or ""),
                        "album": str(r[ix["앨범제목"]] or ""),
                        "genre": str(r[ix["장르"]] or ""),
                    }
                )
    return out


printed = Counter((r["code"], norm(r["album"])) for r in load(S + "/printed.xlsx"))
cur = load(V3)
have = Counter((r["code"], norm(r["album"])) for r in cur)
info = {r["code"]: r for r in cur}
ORDER = [
    "좌측1층",
    "좌측2층",
    "좌측3층",
    "중앙좌3층",
    "중앙좌2층",
    "중앙좌1층",
    "중앙우3층",
    "중앙우2층",
    "중앙우1층",
    "침대옆",
    "우측1층",
    "우측2층",
    "보관함",
]
out = []
for (code, alb), n in have.items():
    need = n - min(n, printed.get((code, alb), 0))
    if need <= 0:
        continue
    r = info[code]
    s, sh, no = secmap.get(code, ("?", "?", 0))
    out.append((s, code, need, sh, no, r["genre"], r["artist"], r["album"], ""))
out.sort(key=lambda x: (ORDER.index(x[3]) if x[3] in ORDER else 99, x[4]))
out = [(i,) + o for i, o in enumerate(out, 1)]
mk(
    "라벨재출력",
    out,
    [
        "순번",
        "섹션",
        "출력할 분류코드",
        "매수",
        "배치 선반",
        "선반 내 순번",
        "장르",
        "아티스트",
        "앨범제목",
        "완료✓",
    ],
    [5, 26, 18, 5, 11, 7, 10, 22, 34, 7],
    [
        f"LP 라벨 재출력 (2026-08-02) — 최초 출력분 대비 코드가 바뀐 것 전부 · "
        f"{len(out)}종 {sum(o[3] for o in out)}장",
        "※ 라벨은 '섹션명 / 분류코드' 두 줄로 인쇄한다(라벨인쇄-LP 시트와 동일 형식).",
        "※ 배치 선반 → 선반 내 순번 순서로 정렬. 한 번 걸어가며 교체하면 된다.",
        "※ 라벨재발급1·2·3·-v3 은 이 한 장으로 대체된다.",
        "※ CD는 라벨 미부착 방침이라 제외.",
    ],
    "C00000",
    1,
)
wb.save(V3)
print(
    f"\n섹션 {len(shelfsec)}개 · 라벨재출력 {len(out)}종 {sum(o[3] for o in out)}장 · 이동 {mv}장"
)
