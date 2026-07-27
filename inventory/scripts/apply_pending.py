#!/usr/bin/env python3
"""ser8 웹UI에서 발급된 신규 음반(pending_albums.json)을 엑셀에 반영한다.

흐름: ser8 `/add` → 분류번호 발급 + pending 큐 적재 → (백엔드에서) 이 스크립트 →
엑셀 위치 시트에 행 추가 → `pipeline.py` 재생성 → md/카탈로그까지 `music-wiki update`.

- 멱등: 이미 엑셀에 있는 (아티스트, 앨범) 조합은 건너뛰고 큐에서 제거.
- 반영 성공분만 큐에서 비움(실패분은 남겨 다음 실행에서 재시도).
- `--dry-run` 으로 무엇이 반영될지 미리 확인 가능.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import shutil
from datetime import datetime

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
INV = os.path.dirname(HERE)
XLSX = "/home/neotango/lpcd_image/20260726_LPCD_목록_v2.xlsx"
if not os.path.exists(XLSX):
    XLSX = os.path.join(INV, "20260726_LPCD_목록_v2.xlsx")
PENDING = os.path.join(INV, "data", "pending_albums.json")
ARCHIVE = os.path.join(INV, "data", "applied_albums.json")
SUMMARY_SHEETS = {"분류코드표", "중복목록", "미식별목록", "정리계획", "라벨인쇄"}

# pending 항목 키 → 엑셀 헤더
FIELD_MAP = {"medium": "매체", "genre": "장르", "artist": "아티스트", "album": "앨범제목",
             "composer": "작곡가", "performer": "연주자", "label_cat": "레이블/카탈로그번호",
             "code": "분류코드", "notes": "비고"}


def norm(s):
    return re.sub(r"[^a-z0-9가-힣]", "", str(s or "").lower())


def existing_keys(wb):
    keys = set()
    for name in wb.sheetnames:
        if name in SUMMARY_SHEETS:
            continue
        ws = wb[name]
        hdr = [str(c.value or "") for c in ws[1]]
        if "아티스트" not in hdr or "앨범제목" not in hdr:
            continue
        ia, ial = hdr.index("아티스트"), hdr.index("앨범제목")
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r and r[ia] and r[ial]:
                keys.add((norm(r[ia]), norm(r[ial])))
    return keys


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not os.path.exists(PENDING):
        print("· 대기 중인 신규 음반 없음")
        return
    pend = json.load(open(PENDING, encoding="utf-8"))
    if not pend:
        print("· 대기 중인 신규 음반 없음")
        return

    wb = openpyxl.load_workbook(XLSX)
    have = existing_keys(wb)
    added, skipped, failed = [], [], []

    for item in pend:
        artist, album = str(item.get("artist") or ""), str(item.get("album") or "")
        key = (norm(artist), norm(album))
        if not (artist.strip() and album.strip()):
            failed.append((item, "아티스트/앨범 누락"))
            continue
        if key in have:
            skipped.append(item)          # 이미 반영됨 → 큐에서 제거
            continue
        sheet = item.get("sheet") or "LP중앙선반2열1층"
        if sheet not in wb.sheetnames:
            failed.append((item, f"시트 없음: {sheet}"))
            continue
        ws = wb[sheet]
        hdr = [str(c.value or "") for c in ws[1]]
        vals = [None] * len(hdr)
        if "순번" in hdr:
            vals[hdr.index("순번")] = ws.max_row      # 헤더 제외 행 수 = 다음 순번
        for k, h in FIELD_MAP.items():
            if h in hdr and item.get(k):
                vals[hdr.index(h)] = item[k]
        if "비고" in hdr:
            note = str(item.get("notes") or "").strip()
            stamp = f"[웹등록 {datetime.now().strftime('%Y-%m-%d')}]"
            vals[hdr.index("비고")] = (note + " " + stamp).strip()
        if not args.dry_run:
            ws.append(vals)
        have.add(key)
        added.append(item)

    print(f"· 대기 {len(pend)}건 → 추가 {len(added)} / 이미있음 {len(skipped)} / 실패 {len(failed)}")
    for it in added:
        print(f"    + {it.get('code', '?'):<16} {it.get('artist')} — {it.get('album')}"
              f"  @{it.get('sheet')}")
    for it, why in failed:
        print(f"    ! {it.get('artist')} — {it.get('album')}: {why}")

    if args.dry_run:
        print("· dry-run — 파일 미변경")
        return
    if added:
        shutil.copy(XLSX, XLSX + ".bak")     # 안전 백업(직전 상태)
        wb.save(XLSX)
        print(f"· 엑셀 저장: {XLSX} (백업 .bak)")

    # 반영·중복분은 큐에서 제거하고 이력에 축적, 실패분만 남긴다
    done = added + skipped
    if done:
        hist = json.load(open(ARCHIVE, encoding="utf-8")) if os.path.exists(ARCHIVE) else []
        stamp = datetime.now().isoformat(timespec="seconds")
        hist += [{**it, "applied_at": stamp} for it in done]
        json.dump(hist, open(ARCHIVE, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    rest = [it for it, _ in failed]
    if rest:
        json.dump(rest, open(PENDING, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    elif os.path.exists(PENDING):
        os.remove(PENDING)
    print(f"· 큐 정리: 남은 {len(rest)}건" + ("" if rest else " (비움)"))
    if added:
        print("· 다음 단계: python inventory/scripts/pipeline.py && music-wiki update")


if __name__ == "__main__":
    main()
