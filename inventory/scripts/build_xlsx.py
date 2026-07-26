#!/usr/bin/env python3
"""Build the LP/CD inventory .xlsx from the per-photo JSON files.

Re-runnable: reads every <scratch>/lpcd/*.json and writes one sheet per photo,
plus a 요약 index sheet. Columns can be changed here and regenerated without
re-recognizing the photos.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRATCH = Path("/home/neotango/lpcd_image/_data")
OUT = Path("/home/neotango/lpcd_image/LPCD_목록.xlsx")

# physical browse order; any file not listed is appended alphabetically
PREFERRED = [
    "CD윗층", "CD아래층",
    "LP중앙선반1열1층", "LP중앙선반1열2층", "LP중앙선반1열3층",
    "LP중앙선반2열1층", "LP중앙선반2열2층", "LP중앙선반2열3층",
    "LP좌측선반1층", "LP좌측선반2층", "LP좌측선반3층",
    "LP우측선반1층", "LP우측선반2층",
    "LP보관박스1", "LP보관박스2", "LP침대옆",
]

HEADERS = ["순번", "매체", "장르", "아티스트", "앨범제목", "레이블/카탈로그번호",
           "Discogs ID", "MusicBrainz MBID", "DB링크", "DB매칭(확인용)", "해설", "비고"]
KEYS = ["order", "medium", "genre", "artist", "album", "label_cat",
        "discogs_id", "mbid", "db_url", "db_match", "desc", "notes"]
WIDTHS = [6, 6, 10, 26, 34, 22, 14, 26, 40, 30, 50, 26]

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=12)


def load() -> dict[str, dict]:
    data: dict[str, dict] = {}
    for f in SCRATCH.glob("*.json"):
        try:
            data[f.stem] = json.loads(f.read_text(encoding="utf-8"))
        except Exception as e:  # noqa: BLE001
            print(f"  ! skip {f.name}: {e}")
    return data


def ordered_names(present: set[str]) -> list[str]:
    names = [n for n in PREFERRED if n in present]
    names += sorted(present - set(names))
    return names


def style_header(ws, row_idx: int, ncols: int = len(HEADERS)) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_sheet(wb: Workbook, name: str, payload: dict) -> tuple[int, int]:
    ws = wb.create_sheet(title=name[:31])
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # title row
    ws.cell(row=1, column=1, value=f"{name}  ({payload.get('medium', '')})").font = TITLE_FONT
    ws.append([])  # row 2 spacer
    ws.append(HEADERS)  # row 3
    style_header(ws, 3)
    rows = payload.get("rows", []) or []
    identified = 0
    for r in rows:
        vals = [r.get(k, "") for k in KEYS]
        if (r.get("artist") or "").strip() or (r.get("album") or "").strip():
            identified += 1
        ws.append(vals)
    ws.freeze_panes = "A4"
    last = 3 + len(rows)
    if len(rows):
        ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}{last}"
    # wrap notes/album columns
    for row in ws.iter_rows(min_row=4, max_row=last, min_col=5, max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    return len(rows), identified


def main() -> None:
    data = load()
    if not data:
        print("no JSON files yet — nothing to build")
        return
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    summary = wb.create_sheet(title="요약")
    summary.append(["사진(시트)", "매체", "항목수", "아티스트/앨범 식별", "인식률"])
    style_header(summary, 1, ncols=5)
    for w, c in zip([22, 8, 8, 16, 10], range(1, 6)):
        summary.column_dimensions[get_column_letter(c)].width = w

    total = 0
    total_id = 0
    for name in ordered_names(set(data)):
        n, ident = write_sheet(wb, name, data[name])
        total += n
        total_id += ident
        rate = f"{(ident / n * 100):.0f}%" if n else "-"
        summary.append([name, data[name].get("medium", ""), n, ident, rate])
    summary.append([])
    summary.append(["합계", "", total, total_id,
                    f"{(total_id / total * 100):.0f}%" if total else "-"])
    summary.cell(row=summary.max_row, column=1).font = Font(bold=True)
    summary.freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  sheets: {len(data)}  items: {total}  identified: {total_id}")


if __name__ == "__main__":
    main()
